import openpyxl
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.user import User, UserRole, AuthProvider
from app.core.security import hash_password
import logging

logger = logging.getLogger(__name__)


async def import_students_from_excel(
    file_bytes: bytes,
    db: AsyncSession,
    default_class: str | None = None,
) -> dict:
    """
    Parse an Excel file and create student accounts.
    Expected columns (row 1 = headers):
      username | full_name | email | password | class_name (optional)

    Returns {"created": N, "skipped": N, "errors": [...]}
    """
    created = 0
    skipped = 0
    errors = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        return {"created": 0, "skipped": 0, "errors": [f"Could not read file: {e}"]}

    # Read header row
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    required = {"username", "full_name", "email", "password"}
    missing = required - set(headers)
    if missing:
        return {
            "created": 0,
            "skipped": 0,
            "errors": [f"Missing required columns: {', '.join(missing)}"]
        }

    col = {name: idx for idx, name in enumerate(headers)}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # skip empty rows

        try:
            username = str(row[col["username"]] or "").strip()
            full_name = str(row[col["full_name"]] or "").strip()
            email = str(row[col["email"]] or "").strip()
            password = str(row[col["password"]] or "").strip()
            class_name = str(row[col["class_name"]] or default_class or "").strip() \
                if "class_name" in col else (default_class or "")

            if not all([username, full_name, email, password]):
                errors.append(f"Row {row_num}: missing required fields")
                skipped += 1
                continue

            # Check if user already exists
            existing = await db.execute(
                select(User).where(
                    (User.username == username) | (User.email == email)
                )
            )
            if existing.scalar_one_or_none():
                logger.info(f"Skipping existing user: {username}")
                skipped += 1
                continue

            user = User(
                username=username,
                email=email,
                full_name=full_name,
                hashed_password=hash_password(password),
                role=UserRole.student,
                auth_provider=AuthProvider.local,
                class_name=class_name or None,
                is_active=True,
            )
            db.add(user)
            created += 1
            logger.info(f"Created user: {username}")

        except Exception as e:
            errors.append(f"Row {row_num}: {e}")
            skipped += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
