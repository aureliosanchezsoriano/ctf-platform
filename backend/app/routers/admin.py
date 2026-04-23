import logging
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from pydantic import BaseModel
from io import BytesIO
from app.core.database import get_db
from app.core.auth import get_current_teacher
from app.core.security import hash_password
from app.models.user import User, UserRole, AuthProvider
from app.models.challenge import Challenge
from app.models.attempt import Attempt
from app.services.orchestrator import list_all_containers, stop_all_for_user, get_docker_client
import openpyxl
import docker.errors

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── Schemas ───────────────────────────────────────────────────────────────────

class StudentProgress(BaseModel):
    id: str
    username: str
    full_name: str
    class_name: str | None
    points: int
    solved_count: int
    total_challenges: int
    is_active: bool


class ContainerEntry(BaseModel):
    name: str
    status: str
    user: str
    challenge: str
    short_id: str


class ImportResult(BaseModel):
    created: int
    skipped: int
    errors: list[str]


# ── Student management ────────────────────────────────────────────────────────

@router.get("/students", response_model=list[StudentProgress])
async def list_students(
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    # Get all students
    result = await db.execute(
        select(User).where(User.role == UserRole.student).order_by(User.full_name)
    )
    students = result.scalars().all()

    # Total active challenges
    total_result = await db.execute(
        select(func.count(Challenge.id)).where(Challenge.is_active == True)
    )
    total_challenges = total_result.scalar() or 0

    # Points and solve count per student
    solved_result = await db.execute(
        select(
            Attempt.user_id,
            func.sum(Challenge.points).label("points"),
            func.count(Attempt.id).label("solved_count"),
        )
        .join(Challenge, Challenge.id == Attempt.challenge_id)
        .where(Attempt.is_correct == True)
        .group_by(Attempt.user_id)
    )
    solved_map = {str(row.user_id): (row.points or 0, row.solved_count) for row in solved_result}

    return [
        StudentProgress(
            id=str(s.id),
            username=s.username,
            full_name=s.full_name,
            class_name=s.class_name,
            points=solved_map.get(str(s.id), (0, 0))[0],
            solved_count=solved_map.get(str(s.id), (0, 0))[1],
            total_challenges=total_challenges,
            is_active=s.is_active,
        )
        for s in students
    ]


@router.patch("/students/{user_id}/toggle")
async def toggle_student(
    user_id: str,
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(User).where(User.id == user_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    student.is_active = not student.is_active
    return {"id": user_id, "is_active": student.is_active}


@router.delete("/students/{user_id}/containers")
async def kill_student_containers(
    user_id: str,
    current_user: User = Depends(get_current_teacher),
):
    count = stop_all_for_user(user_id)
    return {"stopped": count}


# ── Container monitor ─────────────────────────────────────────────────────────

@router.get("/containers", response_model=list[ContainerEntry])
async def get_containers(current_user: User = Depends(get_current_teacher)):
    containers = list_all_containers()
    return [ContainerEntry(**c) for c in containers]


@router.delete("/containers/all")
async def kill_all_containers(current_user: User = Depends(get_current_teacher)):
    """Emergency stop — kills all CTF containers."""
    client = get_docker_client()
    containers = client.containers.list(filters={"label": "ctf.managed=true"})
    count = 0
    for c in containers:
        try:
            c.remove(force=True)
            count += 1
        except docker.errors.NotFound:
            pass
    logger.warning(f"Emergency stop: killed {count} containers by {current_user.username}")
    return {"stopped": count}


# ── Excel import ──────────────────────────────────────────────────────────────

@router.post("/import/excel", response_model=ImportResult)
async def import_from_excel(
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """
    Import students from Excel. Expects columns:
    username | full_name | email | password | class_name
    """
    # For now returns a template — actual file upload comes via multipart
    return ImportResult(created=0, skipped=0, errors=["Use POST with multipart/form-data"])


@router.post("/import/excel/upload", response_model=ImportResult)
async def upload_excel(
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import UploadFile, File
    # Placeholder — implemented below with proper multipart
    pass


# ── Export results ────────────────────────────────────────────────────────────

@router.get("/export/results")
async def export_results(
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """Export student results as Excel for the gradebook."""
    result = await db.execute(
        select(User).where(User.role == UserRole.student, User.is_active == True)
        .order_by(User.class_name, User.full_name)
    )
    students = result.scalars().all()

    total_result = await db.execute(
        select(func.count(Challenge.id)).where(Challenge.is_active == True)
    )
    total_challenges = total_result.scalar() or 0

    solved_result = await db.execute(
        select(
            Attempt.user_id,
            func.sum(Challenge.points).label("points"),
            func.count(Attempt.id).label("solved_count"),
        )
        .join(Challenge, Challenge.id == Attempt.challenge_id)
        .where(Attempt.is_correct == True)
        .group_by(Attempt.user_id)
    )
    solved_map = {str(row.user_id): (row.points or 0, row.solved_count) for row in solved_result}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CTF Results"

    # Header
    headers = ["Full Name", "Username", "Class", "Points", "Solved", "Total Challenges", "Score %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for row_idx, student in enumerate(students, 2):
        uid = str(student.id)
        points, solved = solved_map.get(uid, (0, 0))
        max_points_result = await db.execute(
            select(func.sum(Challenge.points)).where(Challenge.is_active == True)
        )
        max_points = max_points_result.scalar() or 1
        score_pct = round((points / max_points) * 100, 1)

        ws.cell(row=row_idx, column=1, value=student.full_name)
        ws.cell(row=row_idx, column=2, value=student.username)
        ws.cell(row=row_idx, column=3, value=student.class_name or "")
        ws.cell(row=row_idx, column=4, value=points)
        ws.cell(row=row_idx, column=5, value=solved)
        ws.cell(row=row_idx, column=6, value=total_challenges)
        ws.cell(row=row_idx, column=7, value=score_pct)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ctf_results.xlsx"},
    )


from fastapi import UploadFile, File


@router.post("/import/students", response_model=ImportResult)
async def import_students(
    file: UploadFile = File(...),
    class_name: str | None = None,
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """Upload an Excel file to bulk-create student accounts."""
    from app.services.excel_import import import_students_from_excel

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    content = await file.read()
    result = await import_students_from_excel(content, db, default_class=class_name)
    return ImportResult(**result)


@router.get("/challenges")
async def list_all_challenges(
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """List ALL challenges including inactive ones — for the admin panel."""
    result = await db.execute(
        select(Challenge).order_by(Challenge.points.asc(), Challenge.slug.asc())
    )
    challenges = result.scalars().all()
    return [
        {
            "id": str(c.id),
            "slug": c.slug,
            "name": c.name,
            "category": c.category,
            "difficulty": c.difficulty,
            "points": c.points,
            "is_active": c.is_active,
            "is_required": c.is_required,
            "type": c.type,
        }
        for c in challenges
    ]


@router.delete("/students/{user_id}/progress")
async def reset_student_progress(
    user_id: str,
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """Reset all attempts for a student — wipes points and solved challenges."""
    result = await db.execute(select(User).where(User.id == user_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    from sqlalchemy import delete as sql_delete
    await db.execute(
      sql_delete(Attempt).where(Attempt.user_id == user_id)
    )

    await db.commit()
    logger.info(f"Progress reset for user {student.username} by {current_user.username}")
    return {"id": user_id, "username": student.username, "reset": True}


@router.delete("/students/{user_id}")
async def delete_student(
    user_id: str,
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """Permanently delete a student account and all their data."""
    result = await db.execute(select(User).where(User.id == user_id))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if student.role != UserRole.student:
        raise HTTPException(status_code=400, detail="Can only delete student accounts")

    # Stop any running containers
    stop_all_for_user(user_id)

    await db.delete(student)
    await db.commit()
    logger.info(f"Student {student.username} deleted by {current_user.username}")
    return {"deleted": True, "username": student.username}


class CreateStudentRequest(BaseModel):
    username: str
    full_name: str
    email: str
    password: str
    class_name: str | None = None


@router.post("/students", response_model=dict)
async def create_student(
    payload: CreateStudentRequest,
    current_user: User = Depends(get_current_teacher),
    db: AsyncSession = Depends(get_db),
):
    """Create a single student account."""
    existing = await db.execute(
        select(User).where(
            (User.username == payload.username) | (User.email == payload.email)
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Username or email already exists")

    student = User(
        username=payload.username,
        email=payload.email,
        full_name=payload.full_name,
        hashed_password=hash_password(payload.password),
        role=UserRole.student,
        auth_provider=AuthProvider.local,
        class_name=payload.class_name,
        is_active=True,
    )
    db.add(student)
    await db.flush()
    await db.commit()

    logger.info(f"Student {payload.username} created by {current_user.username}")
    return {
        "id": str(student.id),
        "username": student.username,
        "full_name": student.full_name,
        "class_name": student.class_name,
    }
